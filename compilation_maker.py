import os
import re
import io
import time
import random
import pickle
import platform
import subprocess
from datetime import datetime

import openpyxl
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

try:
    from PIL import Image
    import numpy as np
    SHARPNESS_AVAILABLE = True
except Exception:
    SHARPNESS_AVAILABLE = False


# ============================================================
# ============================================================
#                       CONFIG
# ============================================================
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- Audio ---
BGM_VOLUME = 0.25              
ORIGINAL_AUDIO_VOLUME = 1.0    
BGM_FADE_IN_SEC = 1            
BGM_FADE_OUT_SEC = 3           
AUDIO_LIMITER_LEVEL = 0.95     
LOUDNORM_FILTER = "loudnorm=I=-16:TP=-1.5:LRA=11"  

# --- Video length / usage logic ---
TARGET_DURATION_SEC = 480      # Videos will stop adding clips AFTER crossing this threshold
MIN_DURATION_SEC = 240         # Output video will abort if total duration is less than 4 minutes
MAX_USAGE_COUNT = 1            # If a video reaches this count in the logs, it will not be used
VIDEO_BITRATE = "6000k"

# Dynamically set based on user input
OUTPUT_RESOLUTION = (1920, 1080)
IS_VERTICAL = False 

# --- Folders / files ---
DOWNLOAD_DIR = os.path.join(BASE_DIR, "temp_videos")
OUTPUT_DIR = os.path.join(BASE_DIR, "output long form")
TEMP_THUMBS_DIR = os.path.join(BASE_DIR, "temp_thumbs")

# [UPDATE THESE PATHS FOR YOUR LOCAL ENVIRONMENT]
BGM_FOLDER = "[INSERT_YOUR_BGM_FOLDER_PATH_HERE]"
EXCEL_PATH = "[INSERT_YOUR_EXCEL_FILE_PATH_HERE]"

# --- Google Drive ---
SCOPES = ['https://www.googleapis.com/auth/drive']
TOKEN_PATH = os.path.join(BASE_DIR, "token.pickle")
CLIENT_SECRET_PATH = os.path.join(BASE_DIR, "client_secret.json")


# ============================================================
# EXCEL INIT & DATA HANDLING
# ============================================================
def init_excel():
    if not os.path.exists(EXCEL_PATH):
        os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
        wb = openpyxl.Workbook()
        ws_input = wb.active
        ws_input.title = "input"
        ws_input.append(["channel name", "filename", "link"])
        
        ws_logs = wb.create_sheet("logs")
        ws_logs.append(["filename", "count", "id"])
        
        ws_temp = wb.create_sheet("temp")
        ws_temp.append(["channel name", "filename", "link"])
        wb.save(EXCEL_PATH)
    else:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        changed = False
        if "input" not in wb.sheetnames:
            ws = wb.create_sheet("input")
            ws.append(["channel name", "filename", "link"])
            changed = True
        if "logs" not in wb.sheetnames:
            ws = wb.create_sheet("logs")
            ws.append(["filename", "count", "id"])
            changed = True
        if "temp" not in wb.sheetnames:
            ws = wb.create_sheet("temp")
            ws.append(["channel name", "filename", "link"])
            changed = True
        if changed:
            wb.save(EXCEL_PATH)

def load_logs():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["logs"]
    logs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        filename = str(row[0]).strip()
        count = int(row[1]) if len(row) > 1 and row[1] else 0
        file_id = str(row[2]).strip() if len(row) > 2 and row[2] else filename
        logs[file_id] = {"name": filename, "count": count}
    return logs

def finalize_excel_post_run(used_files, usage_log):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        
        ws_logs = wb["logs"]
        ws_logs.delete_rows(2, ws_logs.max_row)
        for file_id, data in usage_log.items():
            ws_logs.append([data["name"], data["count"], file_id])
            
        ws_temp = wb["temp"]
        ws_temp.delete_rows(2, ws_temp.max_row)
        for f in used_files:
            ws_temp.append([f.get("channel", ""), f.get("name", ""), f.get("link", ""), f.get("id", "")])
            
        used_ids = {str(f.get("id")) for f in used_files if f.get("id")}
        ws_input = wb["input"]
        
        rows_to_keep = []
        for row in ws_input.iter_rows(min_row=2, values_only=True):
            if not row or not row[2]: 
                continue
            fid = extract_drive_id(str(row[2]).strip())
            if str(fid) not in used_ids:
                rows_to_keep.append(row)
                
        ws_input.delete_rows(2, ws_input.max_row)
        for row_data in rows_to_keep:
            ws_input.append(row_data)
        
        wb.save(EXCEL_PATH)
        print(f"\n✅ Excel updated: 'logs' saved, 'temp' populated, 'input' updated (used files removed).")
    except Exception as e:
        print(f"❌ Failed to update Excel post-run: {e}")

def extract_drive_id(link):
    patterns = [r"/d/([a-zA-Z0-9_-]{10,})", r"[?&]id=([a-zA-Z0-9_-]{10,})"]
    for p in patterns:
        m = re.search(p, str(link))
        if m: return m.group(1)
    return None

def get_source_sheet_files(target_channel_name):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["input"]
    files = []
    target = target_channel_name.strip().lower()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]: continue
        c_name = str(row[0]).strip() if row[0] else ""
        if c_name.lower() != target:
            continue
        filename = str(row[1]).strip() if row[1] and str(row[1]).strip() else None
        link = str(row[2]).strip()
                
        fid = extract_drive_id(link)
        if not fid: continue
        files.append({"id": fid, "link": link, "name": filename, "channel": c_name})
    return files


# ============================================================
# GOOGLE DRIVE
# ============================================================
def get_drive_service():
    creds = None
    if os.path.exists(TOKEN_PATH):
        with open(TOKEN_PATH, "rb") as token: creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token: creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_PATH, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_PATH, "wb") as token: pickle.dump(creds, token)
    return build('drive', 'v3', credentials=creds)

def _drive_duration_seconds(meta):
    vmm = meta.get("videoMediaMetadata") if meta else None
    if not vmm or vmm.get("durationMillis") is None: return None
    try: return float(vmm["durationMillis"]) / 1000.0
    except (TypeError, ValueError): return None

def download_video(drive_service, file):
    import requests 
    path = os.path.join(DOWNLOAD_DIR, f"{file['id']}.mp4")
    try:
        req = drive_service.files().get_media(fileId=file["id"], supportsAllDrives=True)
        fh = io.FileIO(path, 'wb')
        dl = MediaIoBaseDownload(fh, req)
        done = False
        while not done: dl.next_chunk()
            
        if os.path.getsize(path) == 0:
            print(f"⚠️ File '{file.get('name', file['id'])}' is 0 bytes on Drive. Skipping.")
            safe_remove(path)
            return None
            
        return path
        
    except Exception as e:
        if "416" in str(e):
            print(f"⚠️ Chunked download failed (416) for '{file.get('name', file['id'])}'. Attempting direct stream fallback...")
            try:
                with open(TOKEN_PATH, "rb") as token: creds = pickle.load(token)
                if not creds.valid: 
                    from google.auth.transport.requests import Request as AuthRequest
                    creds.refresh(AuthRequest())
                url = f"https://www.googleapis.com/drive/v3/files/{file['id']}?alt=media&supportsAllDrives=true"
                headers = {'Authorization': f'Bearer {creds.token}'}
                with requests.get(url, headers=headers, stream=True) as r:
                    r.raise_for_status()
                    with open(path, 'wb') as f:
                        for chunk in r.iter_content(chunk_size=1024*1024):
                            if chunk: f.write(chunk)
                if os.path.getsize(path) == 0:
                    print(f"⚠️ Direct stream also returned 0 bytes. The file is empty/corrupted on Google Drive.")
                    safe_remove(path)
                    return None
                print(f"   ✅ Direct stream fallback successful for '{file.get('name', file['id'])}'")
                return path
            except Exception as fallback_e:
                print(f"⚠️ Fallback download also failed: {fallback_e}")
                safe_remove(path)
                return None
        else:
            print(f"⚠️ Download failed for {file.get('name', file.get('id'))}: {e}")
            safe_remove(path)
            return None


# ============================================================
# SELECTION LOGIC
# ============================================================
def get_usage_count(usage_log, file_id):
    return usage_log.get(str(file_id), {}).get("count", 0)

def bump_usage(usage_log, file_id, name):
    fid = str(file_id)
    if fid not in usage_log:
        usage_log[fid] = {"name": name, "count": 0}
    usage_log[fid]["count"] += 1

def select_clips(all_files, usage_log):
    """
    Groups videos strictly by usage count, then heavily randomizes them.
    This guarantees unique, non-sequential picking across the entire sheet.
    """
    groups = {}
    for f in all_files:
        count = get_usage_count(usage_log, f["id"])
        if count >= MAX_USAGE_COUNT:
            print(f"   ⏭ Skipping '{f.get('name', f['id'])}' (Used {count} times, limit is {MAX_USAGE_COUNT})")
            continue
        if count not in groups:
            groups[count] = []
        groups[count].append(f)
        
    pool = []
    for count in sorted(groups.keys()):
        group_files = groups[count]
        random.shuffle(group_files) # Force heavy randomization
        pool.extend(group_files)
        
    return pool


# ============================================================
# UTILS / AUDIO / RENDERING
# ============================================================
def get_duration(path):
    res = subprocess.run(["ffmpeg", "-i", path], stderr=subprocess.PIPE)
    m = re.search(r"Duration: (\d+):(\d+):(\d+\.\d+)", res.stderr.decode(errors="ignore"))
    if m:
        h, mn, s = m.groups()
        return int(h) * 3600 + int(mn) * 60 + float(s)
    return 5

def video_encoder():
    return "h264_videotoolbox" if platform.system() == "Darwin" else "libx264"

def safe_remove(path, label=None):
    if path and os.path.exists(path):
        try: os.remove(path)
        except Exception as e: print(f"⚠️ Could not remove {label or path}: {e}")

def build_bgm_track(target_duration, out_audio_path):
    if not os.path.isdir(BGM_FOLDER): return None
    pool = [f for f in os.listdir(BGM_FOLDER) if f.lower().endswith((".mp3", ".wav", ".m4a", ".aac"))]
    if not pool: return None

    selected, total, working = [], 0.0, []
    while total < target_duration:
        if not working:
            working = pool[:]
            random.shuffle(working)
        pick = working.pop(0)
        path = os.path.join(BGM_FOLDER, pick)
        dur = get_duration(path)
        if dur > 0:
            selected.append(path)
            total += dur

    if not selected: return None
    n = len(selected)
    inputs = []
    for p in selected: inputs += ["-i", p]
    
    concat_labels = "".join(f"[{i}:a]" for i in range(n))
    fade_out_start = max(target_duration - BGM_FADE_OUT_SEC, 0)
    filter_complex = (
        f"{concat_labels}concat=n={n}:v=0:a=1[bgmcat];"
        f"[bgmcat]atrim=0:{target_duration},asetpts=PTS-STARTPTS,"
        f"afade=t=in:d={BGM_FADE_IN_SEC},"
        f"afade=t=out:st={fade_out_start}:d={BGM_FADE_OUT_SEC}[bgm_out]"
    )

    cmd = ["ffmpeg", "-y"] + inputs + ["-filter_complex", filter_complex, "-map", "[bgm_out]", out_audio_path]
    subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return out_audio_path if os.path.exists(out_audio_path) else None

def mix_audio(video_path, bgm_path, out_path, keep_original):
    if keep_original:
        fc = (f"[0:a]volume={ORIGINAL_AUDIO_VOLUME}[a0];[1:a]volume={BGM_VOLUME}[a1];"
              f"[a0][a1]amix=inputs=2:duration=first:normalize=0,alimiter=limit={AUDIO_LIMITER_LEVEL}[aout]")
    else:
        fc = f"[0:a]volume={BGM_VOLUME},alimiter=limit={AUDIO_LIMITER_LEVEL}[aout]"
        
    cmd = ["ffmpeg", "-y", "-i", video_path, "-i", bgm_path, "-filter_complex", fc,
           "-map", "0:v:0", "-map", "[aout]", "-c:v", "copy", "-c:a", "aac", "-shortest", out_path]
    res = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return res.returncode == 0 and os.path.exists(out_path)

def normalize_clip(in_path, out_path):
    """
    Standardizes resolution, framerate, and audio for perfect concatenation.
    This fixes videos freezing, stuttering, or dropping audio.
    """
    probe = subprocess.run(["ffmpeg", "-i", in_path], stderr=subprocess.PIPE)
    stderr_str = probe.stderr.decode('utf-8', errors='ignore')
    has_audio = "Audio:" in stderr_str
    
    enc = video_encoder()
    vf = f"scale={OUTPUT_RESOLUTION[0]}:{OUTPUT_RESOLUTION[1]}:force_original_aspect_ratio=decrease,pad={OUTPUT_RESOLUTION[0]}:{OUTPUT_RESOLUTION[1]}:(ow-iw)/2:(oh-ih)/2,setsar=1,fps=30"
    
    if has_audio:
        cmd = [
            "ffmpeg", "-y", "-i", in_path,
            "-vf", vf, "-c:v", enc, "-b:v", VIDEO_BITRATE, "-pix_fmt", "yuv420p",
            "-c:a", "aac", "-ar", "44100", "-ac", "2", out_path
        ]
    else:
        # Generate an invisible silent audio stream so the final timeline doesn't break
        cmd = [
            "ffmpeg", "-y", "-i", in_path,
            "-f", "lavfi", "-i", "anullsrc=channel_layout=stereo:sample_rate=44100",
            "-vf", vf, "-c:v", enc, "-b:v", VIDEO_BITRATE, "-pix_fmt", "yuv420p",
            "-c:a", "aac", "-shortest", out_path
        ]
        
    res = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if res.returncode != 0 and enc != "libx264":
        cmd[cmd.index(enc)] = "libx264"
        subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        
    return os.path.exists(out_path)

def render_concat(txt_path, out_path, keep_audio):
    """
    Because clips are pre-normalized, we can instantly 'copy' them together without glitching.
    """
    c = ["ffmpeg", "-f", "concat", "-safe", "0", "-i", txt_path]
    c += ["-c:v", "copy"] # Instant passthrough to prevent double-encoding quality loss
    
    if keep_audio:
        c += ["-af", LOUDNORM_FILTER, "-c:a", "aac"]
    else:
        c += ["-an"]
        
    c += ["-y", out_path]
    res = subprocess.run(c, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return res.returncode == 0


# ============================================================
# CORE PIPELINE
# ============================================================
def build_video(channel_name, source_files, drive_service, usage_log):
    """
    Returns a tuple: (used_metadata_list, boolean_was_video_successfully_created)
    """
    display_name = "Compilation" if not channel_name else channel_name
    print(f"\n🎬 Processing Output for: {display_name}")
    
    if not source_files:
        print("❌ No source clips available for this run.")
        return [], False

    ordered = select_clips(source_files, usage_log)
    if not ordered:
        print("❌ All available clips have reached their maximum usage limit.")
        return [], False

    temp_norm_dir = os.path.join(DOWNLOAD_DIR, "normalized")
    os.makedirs(temp_norm_dir, exist_ok=True)

    selected_files = []
    selected_names = set()
    total = 0.0

    print("🔗 Downloading and standardizing clips...")
    for file in ordered:
        if total >= TARGET_DURATION_SEC:
            break
            
        if (file.get("duration") is None or not file.get("name")) and file.get("id"):
            try:
                meta = drive_service.files().get(fileId=file["id"], fields="name,videoMediaMetadata(durationMillis)").execute()
                if not file.get("name"): file["name"] = meta.get("name")
                if file.get("duration") is None: file["duration"] = _drive_duration_seconds(meta)
            except: pass

        if not file.get("name"):
            file["name"] = file["id"]

        fname = file["name"]
        if fname in selected_names: continue

        path = download_video(drive_service, file)
        if not path: continue
        
        print(f"   ⚙️ Normalizing format for '{fname}'...")
        norm_path = os.path.join(temp_norm_dir, f"norm_{file['id']}.mp4")
        
        if not normalize_clip(path, norm_path):
            print(f"   ⚠️ Failed to standardize clip '{fname}'. Skipping.")
            safe_remove(path)
            continue
            
        # Clean up the original messy download to save laptop space
        safe_remove(path) 
        path = norm_path 

        dur = get_duration(path)
        if dur <= 0:
            safe_remove(path)
            continue

        selected_files.append((path, file))
        selected_names.add(fname)
        total += dur

    print(f"\n📊 Added {len(selected_files)} clips | Total duration: {round(total, 2)} sec")

    # If it fails to reach 4 minutes, strictly abort WITHOUT logging the videos
    if total < MIN_DURATION_SEC:
        print(f"❌ Minimum duration limit not met. ({round(total, 2)}s < {MIN_DURATION_SEC}s). Aborting generation. Clips will remain in input sheet for next time.")
        for p, _ in selected_files:
            safe_remove(p) 
        # Return an empty list so it doesn't process them into the final logs
        return [], False

    if not channel_name:
        safe_channel = "Compilation"
    else:
        safe_channel = "".join(channel_name.split())
        
    ts = datetime.now().strftime("%d%m%Y%H%M%S")
    vid_name = f"{safe_channel}{ts}"
    
    vid_folder = os.path.join(OUTPUT_DIR, vid_name)
    os.makedirs(vid_folder, exist_ok=True)

    txt = os.path.join(BASE_DIR, f"inputs_{ts}.txt")
    with open(txt, "w") as f:
        for p, _ in selected_files: f.write(f"file '{p}'\n")

    out = os.path.join(vid_folder, f"{vid_name}.mp4")
    stage = out.replace(".mp4", "_stage.mp4")

    print(f"📁 Output Directory Created: {vid_name}/")
    print("🎞 Concatenating final video...")
    kept_audio = render_concat(txt, stage, True)
    
    # If rendering completely crashes, we also don't log them so we can try again
    if not os.path.exists(stage):
        print("❌ Render failed. Clips will remain in input sheet.")
        for p, _ in selected_files:
            safe_remove(p)
        safe_remove(txt)
        return [], False

    print("🎵 Mixing BGM track...")
    bgm_path = build_bgm_track(get_duration(stage), out.replace(".mp4", "_bgm.m4a"))
    if bgm_path and mix_audio(stage, bgm_path, out, kept_audio):
        safe_remove(stage)
        safe_remove(bgm_path)
    else:
        print("⚠️ Audio mix failed or no BGM found. Outputting video with original audio only.")
        os.replace(stage, out)
        safe_remove(bgm_path)

    for p, _ in selected_files: safe_remove(p)
    safe_remove(txt)

    # SUCCESS: Now we officially log the metadata for removal
    used_metadata = []
    for _, fdata in selected_files:
        bump_usage(usage_log, fdata["id"], fdata.get("name", fdata["id"]))
        used_metadata.append(fdata)

    print(f"✅ Completed Video: {os.path.basename(out)}")
    
    try:
        thumb_path = os.path.join(vid_folder, f"{vid_name}.jpg")
        print("🎨 Generating 16:9 thumbnail...")
        create_thumbnail_from_video(out, thumb_path, TEMP_THUMBS_DIR)
    except Exception as e:
        print(f"⚠️ Thumbnail fail: {e}")

    return used_metadata, True

def start_video_pipeline():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    init_excel()
    
    print("\n📥 Source Mode: Excel Sheet (Inputs tab)")
    channel_input = input("📺 Enter target channel names comma-separated (e.g., ByHand,Mr.Jack,null for blank channels): ").strip()
    
    if not channel_input:
        channels_to_process = ["null"]
    else:
        channels_to_process = [c.strip() for c in channel_input.split(',')]
    
    num_input = input("\n🔢 How many videos to generate per channel? (Leave blank to generate maximum possible): ").strip()
    
    if not num_input:
        num_videos = float('inf')
        print("♾️ Maximum generation selected. Processing until usable clips run out...")
    else:
        try:
            num_videos = int(num_input)
        except ValueError:
            num_videos = 1
            print("⚠️ Invalid number entered, defaulting to 1.")

    drive_service = get_drive_service()
    usage_log = load_logs()
    all_used_files_in_run = []

    for ch in channels_to_process:
        internal_channel_name = "" if ch.lower() == "null" else ch
        
        print(f"\n============================================================")
        print(f"🚀 STARTING PIPELINE FOR: {ch if internal_channel_name else 'null (Blank Channels)'}")
        print(f"============================================================")
        
        generated_count = 0
        while generated_count < num_videos:
            print(f"\n▶ Video Generator: Batch {generated_count+1} for '{ch}'")
            src = get_source_sheet_files(internal_channel_name)
            
            used, success = build_video(internal_channel_name, src, drive_service, usage_log)
            
            # Stricter check: only log if successfully built
            if success and used:
                all_used_files_in_run.extend(used)
                
            if not success:
                print(f"\n🏁 Stopping generation for '{ch}'. (No more usable clips or remaining clips are too short)")
                break
                
            generated_count += 1

    if all_used_files_in_run:
        finalize_excel_post_run(all_used_files_in_run, usage_log)


# ============================================================
# THUMBNAIL LOGIC
# ============================================================
def _extract_frame(video_path, t, out_img):
    subprocess.run(["ffmpeg", "-loglevel", "error", "-ss", str(max(t, 0)), "-i", video_path, "-frames:v", "1", "-y", out_img])
    return os.path.exists(out_img)

def extract_screenshots_from_video(video_path, temp_dir):
    result = subprocess.run(["ffmpeg", "-i", video_path], stderr=subprocess.PIPE, stdout=subprocess.PIPE)
    m = re.search(r"Duration: (\d+):(\d+):(\d+\.\d+)", result.stderr.decode(errors="ignore"))
    if not m: return []
    h, mn, s = m.groups()
    duration = int(h) * 3600 + int(mn) * 60 + float(s)
    
    targets = [min(10, duration * 0.2), duration / 2, max(duration - 20, duration * 0.8)]
    images = []
    for i, t in enumerate(targets):
        out_img = os.path.join(temp_dir, f"thumb_{i}.jpg")
        if _extract_frame(video_path, t, out_img): images.append(out_img)
    return images

def generate_thumbnail_grid(imgs, out_path):
    """
    ALWAYS forces a 16:9 widescreen output grid for the thumbnail,
    even if the source video was rendered in 9:16 vertical format.
    """
    if len(imgs) < 3: return
    
    fc = (
        "[0:v]scale=642:1080:force_original_aspect_ratio=increase,crop=640:1080[v0];"
        "[1:v]scale=642:1080:force_original_aspect_ratio=increase,crop=640:1080[v1];"
        "[2:v]scale=642:1080:force_original_aspect_ratio=increase,crop=640:1080[v2];"
        "[v0][v1][v2]hstack=inputs=3,format=yuvj420p"
    )

    subprocess.run([
        "ffmpeg", "-loglevel", "error", "-i", imgs[0], "-i", imgs[1], "-i", imgs[2],
        "-filter_complex", fc, "-y", out_path
    ])

def create_thumbnail_from_video(video_path, output_path, temp_dir):
    os.makedirs(temp_dir, exist_ok=True)
    imgs = extract_screenshots_from_video(video_path, temp_dir)
    if imgs: generate_thumbnail_grid(imgs, output_path)
    for img in imgs: safe_remove(img)


# ============================================================
# MAIN SETUP PROMPTS
# ============================================================
def set_aspect_ratio():
    global OUTPUT_RESOLUTION, IS_VERTICAL
    print("\n📏 Choose Output Aspect Ratio:")
    print("1. 9:16 (Vertical / Shorts format)")
    print("2. 16:9 (Horizontal / Standard format)")
    choice = input("Enter 1 or 2: ").strip()
    
    if choice == "1":
        OUTPUT_RESOLUTION = (1080, 1920)
        IS_VERTICAL = True
        print("✅ Set to 9:16 (Vertical)")
    else:
        OUTPUT_RESOLUTION = (1920, 1080)
        IS_VERTICAL = False
        print("✅ Set to 16:9 (Horizontal)")


if __name__ == "__main__":
    
    # Prevent macOS from sleeping while the script is running
    if platform.system() == "Darwin":
        print("☕ Caffeinating macOS to prevent sleep while processing...")
        subprocess.Popen(["caffeinate", "-d", "-i", "-m", "-s", "-w", str(os.getpid())])
        
    print("\n🚀 MAIN MENU\n")
    print("1. Only Thumbnails")
    print("2. Video Generation (or Video+Thumbnails pipeline)")
    print("3. Exit")

    choice = input("Choose option: ").strip()

    if choice == "1":
        set_aspect_ratio()
        print("\n🎨 Thumbnail Generator\n")
        vid_path = input("Enter video file path: ").strip().strip("'")
        out_path = input("Enter output jpg path: ").strip().strip("'")
        create_thumbnail_from_video(vid_path, out_path, TEMP_THUMBS_DIR)
    elif choice == "2":
        set_aspect_ratio()
        start_video_pipeline()
    else:
        print("❌ Exiting...")
